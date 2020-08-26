#!/usr/bin/env python

"""
Exports Issues from a repository to an Excel file
Uses basic authentication (Github API Token and Zenhub API Token)
to retrieve Issues from a repository that token has access to.
Supports Github API v3 and ZenHubs current working API.
Derived from https://gist.github.com/Kebiled/7b035d7518fdfd50d07e2a285aff3977
"""
import argparse
import os
import time
import datetime
import requests
import markdown
from retrying import retry
from openpyxl import Workbook
from openpyxl.styles import Font

GITHUB_TOKEN = os.environ['GITHUB_TOKEN']
ZENHUB_TOKEN = os.environ['ZENHUB_TOKEN']
REPO_ID = os.environ['REPO_ID']
ZENHUB_API_RATE_LIMIT = 51

GITHUB_HEADERS = {
    'Accept': 'application/vnd.github.v3+json',
    'Authorization': f'token {GITHUB_TOKEN}',
}

ZENHUB_HEADERS = {'X-Authentication-Token': ZENHUB_TOKEN}

ZEN_URL = f'https://api.zenhub.io/p1/repositories/{REPO_ID}/'

class AttrDict(dict):
    """A dictionary where you can use dot notation for accessing elements."""
    def __getattr__(self, key):
        if key not in self:
            raise AttributeError(key)
        return self[key]

    def __setattr__(self, key, value):
        self[key] = value

    def __delattr__(self, key):
        del self[key]


def get_epics():
    """
    Get the epic(s) on an issue and concatenates them into
    one field.
    :out epics: the response from the call for epics
    """
    zen_url = f'{ZEN_URL}epics/'
    zen_response = requests.get(zen_url, headers=ZENHUB_HEADERS)
    if not zen_response.status_code == 200:
        raise Exception(zen_response.json())
    epics = zen_response.json()
    return epics

def get_dependencies():
    """
    Get the dependencies on all issues
    :out dependencies: the response from the call for dependencies
    """
    zen_url = f'{ZEN_URL}dependencies/'
    print(zen_url)
    zen_response = requests.get(zen_url, headers=ZENHUB_HEADERS)
    if not zen_response.status_code == 200:
        raise Exception(zen_response.json())
    dependencies = zen_response.json()
    return dependencies

def create_blocked_items(issues):
    """
    Create a dictionary for looking up dependenceies by issue number
    :param issues: the dictionary of issues
    :out blocked_items: a dictionary of blocked issues
    """
    response = get_dependencies()
    blocked_items = dict()
    closed = dict()
    fix_blockers = dict()

    for issue in issues:
        if issue['title'].startswith('Fix')and issue['state'] != 'closed':
            #print(f"Do not skip {issue['number']} dependency...")
            fix_blockers[issue['number']] = AttrDict(
                issue_number=issue['number']
            )
    #find all issues to ignore
    for issue in issues:
        if issue['state'] == 'closed':
            milestone = '2200-01-01T23:59:59Z'
        elif issue['milestone']:
            milestone = issue['milestone'].get('due_on', '2200-01-01T23:59:59Z')
        else:
            milestone = '2200-01-01T23:59:59Z'
        if issue['number'] in fix_blockers:
            #print(f"Do not add {issue['number']} to closed...")
            continue
        #if the issue is not due within this sprint or earlier, ignore dependencies
        if datetime.datetime.strptime(milestone, '%Y-%m-%dT%H:%M:%SZ') \
           < datetime.datetime.now():# + datetime.timedelta(days=14):
            closed[issue['number']] = AttrDict(
                issue_number=issue['number']
            )

    for dependency in response['dependencies'] if response['dependencies'] else []:
        issue_number = dependency['blocked']['issue_number']
        blocking = dependency['blocking']['issue_number']
        #if the issue is closed or the dependency is closed, ignore it
        if issue_number in closed or dependency['blocking']['issue_number'] in closed:
            continue
        if issue_number in blocked_items and blocking in fix_blockers:
            temp = dict()
            depends = blocked_items[issue_number].blocked_by
            depends.append(dependency['blocking']['issue_number'])
            temp[issue_number] = AttrDict(
                issue_number=issue_number,
                blocked_by=depends
            )
            blocked_items.update(temp)
        elif blocking in fix_blockers:
            blocked_items[issue_number] = AttrDict(
                issue_number=dependency['blocked']['issue_number'],
                blocked_by=[dependency['blocking']['issue_number']]
            )
    return blocked_items

def get_epic_issues(epic_issue_id):
    """
    Get the epic(s) on an issue and concatenates them into
    one field.
    :param epic_issue_id: the parent epic issue id
    :out epic_issues: a concantenated field with all epics
    """
    epic_url = f'{ZEN_URL}epics/{epic_issue_id}'
    zen_response = get_zenresponse(epic_url)
    epic_issues = zen_response.json()
    return epic_issues

def create_epic_dict():
    """
    Create a dictionary for looking up epics by issue number
    :out issue_epics: a dictionary of issues with the epics it is under
    """
    response = get_epics()
    issue_epics = dict()

    for epic_issue in response['epic_issues'] if response['epic_issues'] else []:
        epic_issue = epic_issue['issue_number']
        epic_issues = get_epic_issues(epic_issue)
        for issue in epic_issues['issues'] if epic_issues['issues'] else []:
            issue_number = issue['issue_number']
            if issue_number in issue_epics:
                temp = dict()
                epic = issue_epics[issue_number].epic_issue
                epic.append(epic_issue)
                temp[issue_number] = AttrDict(
                    issue_number=issue_number,
                    epic_issue=epic
                )
                issue_epics.update(temp)
            else:
                issue_epics[issue_number] = AttrDict(
                    issue_number=issue_number,
                    epic_issue=[epic_issue]
                )
    print('waiting after creating the dictionary')
    time.sleep(45)
    return issue_epics


def get_comments(repo_name, issue_id):
    """
    Get the comments on an issue and concatenates them into
    one field.
    :param repo_name: the name of the github repo
    :param issue_id: the issue number
    :out comment_sum: a concantenated field with all comments
    """
    comments_for_issue_url = f'https://api.github.com/repos/{repo_name}/issues/{issue_id}/comments'
    git_response = requests.get(comments_for_issue_url, headers=GITHUB_HEADERS)
    comments = git_response.json()
    comment_sum = ''
    for comment in comments:
        c_login = comment.get("user", dict()).get('login', "")
        comment_sum = '@'+c_login+' - '+comment_sum + str(comment['body'])
    return comment_sum

def throttle_zenhub(issue_cnt):
    """
    Wait added for the ZenHub api rate limit of 100 requests per minute,
    wait after the rate limit - 1 issues have been processed
    :param issue_cnt: the current issue count
    """
    if issue_cnt%(ZENHUB_API_RATE_LIMIT-1) == 0:
        print(f'{issue_cnt} issues processed')
        time.sleep(45)


def get_assignees(issue):
    """
    Convert the assignees for an issue to a comma-separated string
    :param issue: the issue to find the assignees
    :returns s_assignee_list: the concatenated list of assignees
    """
    s_assignee_list = ''
    for assignee in issue['assignees'] if issue['assignees'] else []:
        s_assignee_list += assignee['login'] + ','
    return s_assignee_list


def get_epics_string(issues, issue):
    """
    Convert the epics for an issue to a comma-separated string
    :param issues: a dictionary mapping of the issues to epics
    :param issue: the current issue
    :returns s_epics: the concatenated list of epic ids
    """
    s_epics = ''
    try:
        epics = issues[issue['number']]['epic_issue']
        for epic in epics if epics else []:
            s_epics += str(epic) + ','
    except KeyError:
        s_epics = ''
    return s_epics


@retry(wait_exponential_multiplier=1000, wait_exponential_max=10000)
def get_zenresponse(issue_url):
    """
    Gets the ZenHub data for the issue
    :param issue_url: the specific url for the issue number
    :returns zen_response: the response for the issue
    """
    zen_response = requests.get(issue_url, headers=ZENHUB_HEADERS)
    if not zen_response.status_code == 200:
        raise Exception(zen_response.json())
    return zen_response


def get_zenhubresponse(issue_number):
    """
    Gets the ZenHub data for the issue
    :param issue_number: the specific issue number
    :returns s_pipeline: the pipeline for the issue
    :returns estimate_value: the estimate for the issue
    """
    zen_url = f'{ZEN_URL}issues/'
    issue_url = f'{zen_url}{issue_number}'
    zen_response = get_zenresponse(issue_url)
    zen_r = zen_response.json()
    s_pipeline = zen_r.get("pipeline", dict()).get('name', "")
    estimate_value = zen_r.get("estimate", dict()).get('value', "")
    return s_pipeline, estimate_value


def get_labels_string(issue):
    """
    concatenates labels into a string
    :param issue: the current issue
    :returns s_labels: the string of comma-separated labels
    :returns s_priority: the priority for the issue
    """
    s_priority = ''
    s_labels = ''
    for label in issue['labels'] if issue['labels'] else []:
        rules = [
            'Low' in label['name'],
            'Medium' in label['name'],
            'High' in label['name'],
        ]
        s_labels += label['name'] + ','
        if any(rules):
            s_priority = label['name']
    return s_labels, s_priority

def calculate_status(issue):
    """
    Calculates the traffic light status on an issue
    :param issue: the issue to review
    :returtns status: the decoded status for the issue
    """
    status = 'Green'

    now = datetime.datetime.now().replace(microsecond=0)

    if issue['state'] == 'closed':
        milestone = '2200-01-01T23:59:59Z'
    elif issue['milestone']:
        milestone = issue['milestone'].get('due_on', now.strftime('%Y-%m-%dT%H:%M:%SZ'))
    else:
        milestone = now.strftime('%Y-%m-%dT%H:%M:%SZ')
    #if the issue is not due within this sprint or earlier, ignore dependencies
    if datetime.datetime.strptime(milestone, '%Y-%m-%dT%H:%M:%SZ') \
       < now + datetime.timedelta(days=14):
        red_rules = [
            datetime.datetime.strptime(milestone, '%Y-%m-%dT%H:%M:%SZ')
            < now,
            issue['blocked'],
        ]
        for label in issue['labels'] if issue['labels'] else []:
            yellow_rules = [
                'yellow' in label['name'],
            ]
            if any(yellow_rules):
                status = 'Yellow'

        if any(red_rules):
            status = 'Red'

    return status

def write_row(issue, row, worksheet):
    """
    Writes rows to an Excel file
    :param issue: the issue to write
    :param row: the row of data to write
    :param worksheet: the worksheet to store the rows
    """
    issue_number = str(issue['number'])
    row['s_pipeline'], row['estimate_value'] = get_zenhubresponse(issue_number)
    status = calculate_status(issue)
    if row['s_state'] == 'closed':
        row['s_pipeline'] = 'Closed'
    row['comments'] = ''
    if issue['comments'] > 0:
        row['comments'] = get_comments(row['repo_name'], issue_number)
    if (not row['s_priority'] and row['s_pipeline'] == 'Closed'):
        row['s_priority'] = 'High'
    rowvalues = [row['repo_name'], issue['number'], issue['title'],
                 row['userstory'], row['s_pipeline'], issue['user']['login'], issue['created_at'],
                 issue['milestone']['title'] if issue['milestone']
                 else "", issue['milestone']['due_on'] if issue['milestone'] else "",
                 row['s_assignee_list'][:-1], row['estimate_value'],
                 row['s_priority'], row['s_labels'],
                 row['comments'], row['s_epics'][:-1], status,
                 issue['blocked'], str(issue['blocked_by']).strip('[]')]
    for i, value in enumerate(rowvalues):
        worksheet.cell(column=(i+1),
                       row=1+row['issue_cnt'],
                       value=value)

def write_headers(worksheet):
    """
    Writes headers to an Excel file
    :param worksheet: the worksheet where the headers are written
    """
    headers = ['Repository', 'Issue Number', 'Issue Title', 'User Story', 'Pipeline',
               'Issue Author', 'Created At', 'Milestone', 'Milestone End Date',
               'Assigned To', 'Estimate Value', 'Priority', 'Labels', 'Comments',
               'Epics', 'Status', 'Blocked', 'Blocked By']
    for i, header in enumerate(headers):
        worksheet.cell(column=(i+1), row=1, value=header)
        worksheet.cell(column=(i+1), row=1).font = Font(bold=True)


def write_issues(issues, args, epics):
    """
    Writes issues to an Excel file
    :param issues: the dictionary of issues
    :param args: the arguments passed in
    :param epics: dictionary of the mapping for issues to epics
    :return issue_cnt: the count of issues processed
    """
    issue_cnt = 0

    repo_name = args.repo[0]
    if args.filename:
        filename = args.filename
    else:
        filename = f"{repo_name.split('/')[1]}.xlsx"
    fileoutput = Workbook()

    worksheet = fileoutput.create_sheet(title="Data")
    defaultsheet = fileoutput['Sheet']
    fileoutput.remove(defaultsheet)

    write_headers(worksheet)

    for issue in issues:
        issue_cnt += 1
        s_labels, s_priority = get_labels_string(issue)
        if args.html == 1:
            userstory = markdown.markdown(issue['body'])
        else:
            userstory = issue['body']
        row = dict(repo_name=repo_name,
                   repo_id=REPO_ID,
                   userstory=userstory,
                   s_assignee_list=get_assignees(issue),
                   s_priority=s_priority,
                   s_labels=s_labels,
                   s_epics=get_epics_string(epics, issue),
                   s_state=issue['state'],
                   issue_cnt=issue_cnt,)
        write_row(issue, row, worksheet)
        print(f'issue count: {issue_cnt}')
        throttle_zenhub(issue_cnt)

    fileoutput.save(filename=filename)
    return issue_cnt


def get_pages(issue_response):
    """
    gets additional pages information
    :param issue_response: the response for the github call for the issue
    :returns pages_dict: a dictionary object for the pages link
    """
    pages_dict = dict(
        (rel[6:-1], url[url.index('<') + 1:-1]) for url, rel in
        [link.split(';') for link in
         issue_response.headers['link'].split(',')])
    return pages_dict


def get_nextpage_response(pages):
    """
    gets the next page information
    :param pages: a dictionary object for the pages link
    :returns issue_response.json(): a json document of the next page link
    """
    issue_response = requests.get(pages['next'], headers=GITHUB_HEADERS)
    if not issue_response.status_code == 200:
        print(issue_response.json())
        raise Exception(issue_response.status_code)
    return issue_response.json()

def get_github_issues(args, state='all'):
    """
    Get github issues
    :param args: the arguments passed in
    :param state: the state of the issue, all by default
    :returns issues: the dictionary of issues
    """

    since = args.since
    repo_name = args.repo[0]
    print(repo_name)

    issues_for_repo_url = f'https://api.github.com/repos/{repo_name}/issues?state={state}'
    issues = []
    if since:
        since_date = datetime.datetime.strptime(since, '%Y-%m-%d').strftime("%Y-%m-%dT%H:%M:%SZ")
        print(f'Filtering since {since_date}...')
        issues_for_repo_url = f'{issues_for_repo_url}&since={since_date}'
        print(f'Request {issues_for_repo_url}...')
    issue_response = requests.get(issues_for_repo_url, headers=GITHUB_HEADERS)
    if not issue_response.status_code == 200:
        raise Exception(issue_response.json())
    response = issue_response.json()
    issues += response
    # more pages? examine the 'link' header returned
    if 'link' in issue_response.headers:
        pages = get_pages(issue_response)
        while 'last' in pages and 'next' in pages:
            pages = get_pages(issue_response)
            issue_response = requests.get(pages['next'], headers=GITHUB_HEADERS)
            response = issue_response.json()
            issues += response
            if pages['next'] == pages['last']:
                blocked = create_blocked_items(issues)
                for issue in issues:
                    if issue['number'] in blocked:
                        issue.update({"blocked": "blocked",
                                      "blocked_by": blocked[issue['number']]['blocked_by'],})
                    else:
                        issue.update({"blocked": "",
                                      "blocked_by": "",})
                    issues[issues.index(issue)] = issue
                return issues
    blocked = create_blocked_items(issues)
    for issue in issues:
        if issue['number'] in blocked:
            issue.update({"blocked": "blocked",
                          "blocked_by": blocked[issue['number']]['blocked_by'],})
        else:
            issue.update({"blocked": "",
                          "blocked_by": "",})
        issues[issues.index(issue)] = issue
    return issues

def get_issues(args):
    """
    Get an issue attributes
    :param args: the arguments passed in
    """

    #get the epic dictionary
    epic_dict = create_epic_dict()
    issues = get_github_issues(args, state=args.state)
    write_issues(issues, args, epic_dict)

def main():
    """The real main function..."""

    parser = argparse.ArgumentParser()
    parser.add_argument('--filename', default=None, help='filename=filename.txt')
    parser.add_argument('--repo', nargs='+', help='repo owner/repo')
    parser.add_argument('--html', default=0, type=int, help='html=1')
    parser.add_argument('--since', default=None, help='since date in the format of 2018-01-01')
    parser.add_argument('--state', default='all', help='the state as defined by github')
    args = parser.parse_args()

    get_issues(args)

if __name__ == '__main__':
    main()
